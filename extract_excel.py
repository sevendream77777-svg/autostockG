import openpyxl
import os

def extract_text_from_excel(excel_file_path, output_txt_path):
    """
    지정된 Excel 파일(.xlsx)의 모든 시트에서 텍스트를 추출하여
    하나의 .txt 파일로 저장합니다.
    """
    
    # 1. 엑셀 파일이 존재하는지 확인
    if not os.path.exists(excel_file_path):
        print(f"오류: 엑셀 파일({excel_file_path})을 찾을 수 없습니다.")
        print("스크립트와 같은 폴더에 엑셀 파일이 있는지 확인하세요.")
        return

    try:
        # 2. 엑셀 파일 열기 (read_only=True로 속도 향상)
        workbook = openpyxl.load_workbook(excel_file_path, read_only=True)
        
        print(f"엑셀 파일 로드 성공: {excel_file_path}")
        
        # 3. 텍스트 파일 열기 (쓰기 모드)
        with open(output_txt_path, 'w', encoding='utf-8') as txt_file:
            
            # 4. 엑셀의 모든 시트(sheet)를 순회
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                print(f" - '{sheet_name}' 시트에서 텍스트 추출 중...")
                
                txt_file.write(f"\n--- [{sheet_name} 시트 시작] ---\n\n")
                
                # 5. 시트의 모든 행(row)을 순회
                for row in sheet.iter_rows():
                    row_texts = []
                    # 6. 행의 모든 셀(cell)을 순회
                    for cell in row:
                        if cell.value is not None:
                            # 셀의 값(텍스트)을 문자열로 변환하여 리스트에 추가
                            row_texts.append(str(cell.value))
                    
                    # 한 줄에 있던 셀 텍스트들을 탭(Tab)으로 구분하여 파일에 쓰기
                    if row_texts:
                        txt_file.write("\t".join(row_texts) + "\n")
                
                txt_file.write(f"\n--- [{sheet_name} 시트 종료] ---\n\n")

        print(f"\n✅ 추출 완료! 모든 텍스트가 {output_txt_path} 파일에 저장되었습니다.")

    except openpyxl.utils.exceptions.InvalidFileException:
        print(f"오류: {excel_file_path} 파일이 올바른 .xlsx 파일이 아닙니다.")
        print("참고: .xls 파일(구버전)은 이 스크립트로 열 수 없습니다.")
    except Exception as e:
        print(f"오류: 엑셀 파일 처리 중 예외 발생: {e}")

# --- 메인 실행 ---
if __name__ == "__main__":
    # --- [설정] ---
    # 1. 키움증권 API 매뉴얼 엑셀 파일의 정확한 이름을 여기에 넣으세요.
    EXCEL_FILE_NAME = "키움 REST API 문서.xlsx" 
    
    # 2. 텍스트를 저장할 출력 파일 이름
    OUTPUT_TXT_NAME = "api_manual_output.txt"
    # --- [설정 끝] ---
    
    # 엑셀 파일이 이 스크립트와 같은 폴더에 있다고 가정
    current_folder = os.path.dirname(__file__)
    excel_path = os.path.join(current_folder, EXCEL_FILE_NAME)
    txt_path = os.path.join(current_folder, OUTPUT_TXT_NAME)
    
    extract_text_from_excel(excel_path, txt_path)