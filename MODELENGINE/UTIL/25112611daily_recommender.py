# ============================================================
# daily_recommender_V34_plus.py  (FINAL FULL VERSION)
# [Update] AI 분석 후 결과 통합 저장 기능 추가
# [Patch] 엑셀 자동 서식 및 통합 리포트 생성 기능 추가
# [FIXED] find_engine_real() - 날짜 형식(4자리/6자리) 비교 오류 수정 및 cands NameError 수정
# [FIXED] load_latest_db() - NameError 수정 (정의 누락 복구)
# ============================================================
import os, sys, argparse, pickle, warnings
import numpy as np
import pandas as pd
from datetime import datetime
import google.generativeai as genai  # Gemini API
import re # [추가]: 정규표현식 사용을 위해 import

# [Patch] 엑셀 서식 관련 라이브러리 추가
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill

# 경로 설정
current_dir = os.path.dirname(os.path.abspath(__file__))
parent_dir  = os.path.dirname(current_dir)   # MODELENGINE
root_dir    = os.path.dirname(parent_dir)    # Root
sys.path.append(root_dir)
try:
    from MODELENGINE.UTIL.config_paths import get_path
    from MODELENGINE.UTIL.version_utils import find_latest_file
except:
    sys.path.append(parent_dir)
    from UTIL.config_paths import get_path
    from UTIL.version_utils import find_latest_file


# ==========================================
# [설정] AI Studio API Key 입력
# ==========================================
def load_api_key():
    """외부 텍스트 파일에서 API 키를 읽어옵니다."""
    key_path = r"C:\공유주방\!개인폴더\!이호정이사\각종key_appkey_decret\googlegemini_api.txt"
    try:
        if os.path.exists(key_path):
            with open(key_path, "r", encoding="utf-8") as f:
                return f.read().strip()
        else:
            print(f"⚠ [Warning] 키 파일을 찾을 수 없습니다: {key_path}")
    except Exception as e:
        print(f"⚠ [Error] 키 파일 읽기 실패: {e}")
    return None

GEMINI_API_KEY = load_api_key()
# ==========================================


# ============================================================
# 유틸 함수들
# ============================================================
def _hash_list(lst):
    """피처 리스트 해시값 생성"""
    s = "|".join(map(str, lst))
    return str(abs(hash(s)))


def pick_close_col(df):
    """Close/ClosePrice/종가/가격 자동 인식"""
    cand = ["Close","close","ClosePrice","종가","가격","Adj Close","AdjClose"]
    for c in cand:
        if c in df.columns:
            return c
    nums = [c for c in df.columns if df[c].dtype.kind in ("i","f")]
    if len(nums) == 1:
        return nums[0]
    raise KeyError("종가 컬럼 찾지 못함(Close/ClosePrice/종가/가격).")



def find_engine_real():
    """
    [수정됨] 엔진 파일 중 데이터날짜 최신 → h=5,w=60,n=1000 기준과 가장 가까운 옵션 선택.
    날짜 형식(4자리/6자리) 통일하여 최신 파일 선택 오류를 수정함.
    """
    base_root = get_path("HOJ_ENGINE")
    if os.path.isfile(base_root):
        base_root = os.path.dirname(base_root)
    real_dir = os.path.join(base_root, "REAL")
    if not os.path.isdir(real_dir):
        raise FileNotFoundError("REAL 폴더 없음: " + real_dir)

    # 후보 수집
    cands = [] # [FIXED] NameError 방지
    for fn in os.listdir(real_dir):
        if fn.startswith("HOJ_ENGINE_REAL") and fn.endswith(".pkl"):
            cands.append(fn)
    if not cands:
        raise FileNotFoundError("REAL 폴더에 엔진이 없습니다.")

    def parse(fn):
        # extract tokens
        parts = fn.split("_")
        date_token = parts[-1].replace(".pkl","")
        
        d = -1
        h=w=n=None

        # 파라미터(h, w, n) 추출 로직 (기존 로직 유지)
        for p in parts:
            if p.startswith("h"):
                try: h=int(p[1:])
                except: pass
            if p.startswith("w"):
                try: w=int(p[1:])
                except: pass
            if p.startswith("n"):
                try: n=int(p[1:])
                except: pass
        
        # [수정된 로직]: 날짜 토큰을 YYMMDD 6자리 숫자로 통일하여 비교 가능하게 함 (25년 기준 가정)
        try:
            if len(date_token) == 6 and date_token.startswith("25"): # YYMMDD (예: 251122)
                d = int(date_token)
            elif len(date_token) == 4: # MMDD (예: 1125). 앞에 '25'를 붙여 YYMMDD로 통일
                d = int("25" + date_token) 
            elif len(date_token) == 8 and date_token.startswith("20"): # YYYYMMDD -> YYMMDD로 변환
                d = int(date_token[2:])
            else:
                 # 파일명 중간에 6자리 날짜가 있으면 그것을 날짜로 사용 (옵션만 붙은 파일 처리 포함)
                 match = re.search(r'(\d{6})\.pkl$', fn)
                 if match and match.group(1).startswith("25"):
                     d = int(match.group(1))
                 elif len(date_token) >= 6 and date_token.startswith("25"): # 숫자로만 이루어진 마지막 토큰이 6자리 이상이고 25로 시작하면 일단 인정
                     d = int(date_token)
                 else:
                    d = -1
        except Exception:
            d = -1
        # [수정된 로직 끝]

        return d,h,w,n

    parsed=[]
    for fn in cands:
        d,h,w,n = parse(fn)
        parsed.append((d,h,w,n,fn))

    # 최신 날짜
    # d=-1 인 파일은 제외하고 최댓값을 찾습니다.
    valid_dates = [p[0] for p in parsed if p[0] != -1]
    if not valid_dates:
        # 이전에 cands는 존재했으므로, 날짜 파싱이 모두 실패했음을 의미
        raise FileNotFoundError("REAL 폴더의 엔진 파일에서 유효한 날짜 태그를 찾을 수 없습니다.")
        
    maxd = max(valid_dates)
    same=[p for p in parsed if p[0]==maxd]

    # 기준값
    H0=5; W0=60; N0=1000

    def score(p):
        _,h,w,n,fn = p
        sh = abs((h or H0)-H0)
        sw = abs((w or W0)-W0)
        sn = abs((n or N0)-N0)
        return (sh, sw, sn, fn)

    chosen = min(same, key=score)
    return os.path.join(real_dir, chosen[4])

def load_latest_db(version="V31"): # [FIXED] NameError 해결을 위해 함수 정의 복구
    """DB 디렉토리에서 최신 통합 DB 파일을 찾아 로드합니다."""
    db_dir = get_path("HOJ_DB")
    latest = find_latest_file(db_dir, f"HOJ_DB_{version}")
    if not latest:
        raise FileNotFoundError("DB를 찾지 못했습니다.")

    df = pd.read_parquet(latest)
    df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
    return df, latest


# ============================================================
# [패치 추가] 엑셀 서식 자동 조정 함수
# ============================================================
def auto_adjust_column_width(worksheet):
    """ 엑셀 컬럼 너비 자동 맞춤 및 헤더 스타일링 """
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter 
        
        # 헤더 스타일 적용
        col[0].font = header_font
        col[0].fill = header_fill
        col[0].alignment = Alignment(horizontal='center')

        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        # 너비 조정
        adjusted_width = (max_length + 2) * 1.1
        adjusted_width = min(adjusted_width, 50) # 최대 50
        adjusted_width = max(adjusted_width, 10) # 최소 10
        worksheet.column_dimensions[column].width = adjusted_width


# ============================================================
# [수정] Gemini 분석 함수 (결과 텍스트 반환하도록 변경)
# ============================================================
def get_gemini_analysis(df):
    """Gemini를 이용해 Top10 분석 텍스트를 생성하여 반환"""
    
    if not GEMINI_API_KEY:
        return "\n[Gemini] API Key가 없어 분석을 생략합니다.\n"

    try:
        genai.configure(api_key=GEMINI_API_KEY)

        # 모델 자동 선택 로직
        all_models = list(genai.list_models())
        valid_model_name = None
        
        # 1순위: Flash (빠름)
        for m in all_models:
            if 'generateContent' in m.supported_generation_methods and 'flash' in m.name:
                valid_model_name = m.name
                break
        
        # 2순위: Pro
        if valid_model_name is None:
            for m in all_models:
                if 'generateContent' in m.supported_generation_methods and 'pro' in m.name:
                    valid_model_name = m.name
                    break
        
        # 기본값
        if valid_model_name is None:
            valid_model_name = "models/gemini-1.5-flash"

        model = genai.GenerativeModel(valid_model_name)
        target_list_str = df.to_string(index=False)
        
        prompt = f"""
아래는 오늘의 HOJ Top10 종목 리스트입니다. 
이 종목들을 기반으로 상승 가능성이 높은 종목을 3개 추천해 주세요.

[Top10 종목 데이터]
{target_list_str}

[요구사항]
- 추천 사유 1~2줄 포함
- 상승 가능성이 높은 순서대로 3개만 제시

[형식]
=== Gemini's Pick ===
1. 종목명: 사유
2. 종목명: 사유
3. 종목명: 사유
"""
        print(f"[Gemini] 모델 '{valid_model_name}' 분석 실행 중...", end="\r") # 진행중 표시
        response = model.generate_content(prompt)
        print(f"[Gemini] 분석 완료.                                 ") # 지우기

        return response.text

    except Exception as e:
        return f"\n⚠ Gemini 분석 중 오류 발생: {e}\n"


# ============================================================
# 메인 로직
# ============================================================
def main(rank_by="combo", topk=10, version="V31"):

    # 1. 엔진 및 DB 로드
    eng_path = find_engine_real()
    with open(eng_path, "rb") as f:
        payload = pickle.load(f)
    model_cls = payload["model_cls"]
    model_reg = payload["model_reg"]
    features  = payload["features"]
    
    df, db_path = load_latest_db(version) # [FIXED] 이제 load_latest_db가 정의되어 오류 발생 안함
    max_date = df["Date"].max()
    df_d = df[df["Date"] == max_date].copy()
    close_col = pick_close_col(df_d)

    # 2. 피처 확인 및 예측
    db_features = [c for c in features if c in df_d.columns]
    X = df_d[db_features].copy()
    mask = X.notnull().all(axis=1)
    df_d = df_d.loc[mask].copy()
    X = X.loc[mask]

    prob = model_cls.predict_proba(X)[:,1]
    ret  = model_reg.predict(X)
    ret_clip = np.clip(ret, -0.10, None)
    combo = prob * ret_clip

    # 3. 결과 DataFrame 생성
    df_out = pd.DataFrame({
        "종목명": df_d.get("Name", df_d.get("name")),
        "종목코드": df_d.get("Code", df_d.get("code")),
        "현재가": df_d[close_col],
        "상승확률(%)": (prob*100).round(2),
        "예측수익률(%)": (ret*100).round(2),
        "동시적용 기대수익(%)": (combo*100).round(2),
    })

    # 정렬
    keymap = {"combo":"동시적용 기대수익(%)", "prob":"상승확률(%)", "ret":"예측수익률(%)"}
    sort_key = keymap.get(rank_by, "동시적용 기대수익(%)")
    df_out = df_out.sort_values(sort_key, ascending=False).head(topk)

    # ------------------------------------------------------------
    # 4. [변경] AI 분석 먼저 실행
    # ------------------------------------------------------------
    ai_result_text = get_gemini_analysis(df_out)

    # ------------------------------------------------------------
    # 5. [변경] 통합 리포트 생성 (출력 및 저장용)
    # ------------------------------------------------------------
    report_content = []
    report_content.append("=" * 60)
    report_content.append(f"📈 HOJ AI Daily Report [{max_date.date()}]")
    report_content.append("=" * 60)
    report_content.append(f"\n[1] 예측 Top {topk} (기준: {rank_by})")
    report_content.append("-" * 60)
    report_content.append(df_out.to_string(index=False))
    report_content.append("-" * 60)
    report_content.append("\n[2] Gemini AI Investment Opinion")
    report_content.append("-" * 60)
    report_content.append(ai_result_text.strip())
    report_content.append("=" * 60)
    
    full_report_str = "\n".join(report_content)

    # ------------------------------------------------------------
    # 6. 화면 출력
    # ------------------------------------------------------------
    print(full_report_str)
    print(f"\n[ENGINE] {os.path.basename(eng_path)}")
    print(f"[DB]     {os.path.basename(db_path)}")

    # ------------------------------------------------------------
    # 7. 파일 저장 (CSV + TXT 리포트)
    # ------------------------------------------------------------
    out_dir = get_path("OUTPUT")
    os.makedirs(out_dir, exist_ok=True)
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    
    # (1) CSV 저장 (데이터용)
    csv_name = f"recommendation_HOJ_V34_{max_date.date()}_{timestamp}_{rank_by}.csv"
    csv_path = os.path.join(out_dir, csv_name)
    df_out.to_csv(csv_path, index=False, encoding="utf-8-sig")

    # (2) TXT 리포트 저장 (보기 편한 용도, AI의견 포함)
    txt_name = f"Report_HOJ_V34_{max_date.date()}_{timestamp}.txt"
    txt_path = os.path.join(out_dir, txt_name)
    
    with open(txt_path, "w", encoding="utf-8") as f:
        f.write(full_report_str)
        f.write(f"\n\n[File Info]\nCSV Data: {csv_name}\nEngine: {os.path.basename(eng_path)}")

    print(f"[SAVE]   CSV: {csv_name}")
    print(f"[SAVE]   TXT: {txt_name} (AI 분석 포함)")

    # ------------------------------------------------------------
    # 8. [패치] 엑셀 리포트 자동 생성 (Format + AI Text)
    # ------------------------------------------------------------
    excel_name = f"Final_Report_HOJ_{max_date.date()}_{timestamp}.xlsx"
    excel_path = os.path.join(out_dir, excel_name)
    
    try:
        print(f"\n[*] Generating Formatted Excel: {excel_name}...")
        
        # (1) Pandas로 데이터 쓰기
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            # Sheet 1: Top 10 추천
            df_out.to_excel(writer, sheet_name='Top 10 추천', index=False)
            
            # Sheet 2: AI 해석 (텍스트)
            df_report = pd.DataFrame({'AI 분석 리포트': [full_report_str]})
            df_report.to_excel(writer, sheet_name='AI 해석', index=False)

        # (2) OpenPyXL로 서식 다듬기
        wb = load_workbook(excel_path)
        
        # Sheet 1 서식 (컬럼 너비 자동, 헤더 스타일)
        if 'Top 10 추천' in wb.sheetnames:
            ws = wb['Top 10 추천']
            auto_adjust_column_width(ws)
            
        # Sheet 2 서식 (줄바꿈, 너비 확장)
        if 'AI 해석' in wb.sheetnames:
            ws = wb['AI 해석']
            cell = ws['A2'] # 본문 셀
            cell.alignment = Alignment(wrap_text=True, vertical='top') # 줄바꿈 허용
            ws.column_dimensions['A'].width = 100 # 넓게 잡기
            
            # 행 높이 늘리기 (내용 길이에 비례)
            line_count = full_report_str.count('\n') + (len(full_report_str) // 100)
            ws.row_dimensions[2].height = max(line_count * 15, 400)

        wb.save(excel_path)
        print(f"[SAVE]   Excel: {excel_name} (서식 적용 완료)")
        
    except Exception as e:
        print(f"[Error] 엑셀 생성 실패: {e}")


# ============================================================
# CLI
# ============================================================
if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--rank_by", default="combo", help="combo | prob | ret")
    ap.add_argument("--topk", type=int, default=10)
    ap.add_argument("--version", default="V31")
    args = ap.parse_args()
    main(rank_by=args.rank_by, topk=args.topk, version=args.version)