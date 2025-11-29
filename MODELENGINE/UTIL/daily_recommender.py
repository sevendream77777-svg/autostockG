# ============================================================
# daily_recommender_V34_plus.py  (FINAL FULL VERSION)
# [Update] AI 분석 후 결과 통합 저장 기능 추가
# [Patch] 엑셀 자동 서식 및 통합 리포트 생성 기능 추가
# [FIXED] find_engine_real() - 날짜 형식(4자리/6자리) 비교 오류 수정 및 cands NameError 수정
# [FIXED] load_latest_db() - NameError 수정 (정의 누락 복구)
# ============================================================
import os, sys, argparse, pickle, warnings
import numpy as np
import pandas as pd
from datetime import datetime
import google.generativeai as genai  # Gemini API
import re # [추가]: 정규표현식 사용을 위해 import
import json
from pathlib import Path

# [Patch] 엑셀 서식 관련 라이브러리 추가
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill

# 경로 설정
current_dir = os.path.dirname(os.path.abspath(__file__))
parent_dir  = os.path.dirname(current_dir)   # MODELENGINE
root_dir    = os.path.dirname(parent_dir)    # Root
sys.path.append(root_dir)
try:
    from MODELENGINE.UTIL.config_paths import get_path
    from MODELENGINE.UTIL.version_utils import find_latest_file
except:
    sys.path.append(parent_dir)
    from UTIL.config_paths import get_path
    from UTIL.version_utils import find_latest_file


# ==========================================
# [설정] AI Studio API Key 입력
# ==========================================
def load_api_key():
    """외부 텍스트 파일에서 API 키를 읽어옵니다."""
    key_path = r"C:\공유주방\!개인폴더\!이호정이사\각종key_appkey_decret\googlegemini_api.txt"
    try:
        if os.path.exists(key_path):
            with open(key_path, "r", encoding="utf-8") as f:
                return f.read().strip()
        else:
            print(f"⚠ [Warning] 키 파일을 찾을 수 없습니다: {key_path}")
    except Exception as e:
        print(f"⚠ [Error] 키 파일 읽기 실패: {e}")
    return None

GEMINI_API_KEY = load_api_key()
# ==========================================


# ============================================================
# 유틸 함수들
# ============================================================
def _hash_list(lst):
    """피처 리스트 해시값 생성"""
    s = "|".join(map(str, lst))
    return str(abs(hash(s)))


def pick_close_col(df):
    """Close/ClosePrice/종가/가격 자동 인식"""
    cand = ["Close","close","ClosePrice","종가","가격","Adj Close","AdjClose"]
    for c in cand:
        if c in df.columns:
            return c
    nums = [c for c in df.columns if df[c].dtype.kind in ("i","f")]
    if len(nums) == 1:
        return nums[0]
    raise KeyError("종가 컬럼 찾지 못함(Close/ClosePrice/종가/가격).")



def find_engine_real():
    """
    [수정] 엔진 파일 선택 로직 (Strict Mode)
    1. 목표 옵션(h, w, n)과 '정확히 일치'하는 파일만 필터링 (타협 없음)
    2. 일치하는 파일들 중 '데이터 날짜(YYMMDD)가 가장 최신'인 파일 선택
    3. 날짜도 같으면 '생성 시간(mtime)'이 최신인 파일 선택
    4. 조건에 맞는 파일이 없으면 명확하게 에러 발생 (FileNotFoundError)
    """
    base_root = get_path("HOJ_ENGINE")
    if os.path.isfile(base_root):
        base_root = os.path.dirname(base_root)
    real_dir = os.path.join(base_root, "REAL")
    if not os.path.isdir(real_dir):
        raise FileNotFoundError("REAL 폴더 없음: " + real_dir)

    cands = []
    for fn in os.listdir(real_dir):
        if fn.startswith("HOJ_ENGINE_REAL") and fn.endswith(".pkl"):
            cands.append(fn)
    if not cands:
        raise FileNotFoundError("REAL 폴더에 엔진 파일이 하나도 없습니다.")

    # ---------------------------------------------------------
    # [설정] 찾고자 하는 엔진의 목표 옵션 (Strict Target)
    # 추후 UI 연동 시 이 부분을 인자로 받게 수정 가능
    # ---------------------------------------------------------
    TARGET_H = 5      # Horizon
    TARGET_W = 60     # Input Window
    TARGET_N = 1000   # Estimators

    valid_candidates = []

    for fn in cands:
        # 1. 날짜 파싱 (YYMMDD 6자리 통일)
        parts = fn.split("_")
        date_token = parts[-1].replace(".pkl","")
        d = -1
        try:
            if len(date_token) == 6 and date_token.startswith("25"):
                d = int(date_token)
            elif len(date_token) == 4:
                d = int("25" + date_token)
            elif len(date_token) == 8 and date_token.startswith("20"):
                d = int(date_token[2:])
            else:
                 # t251125 등 복잡한 패턴 처리
                 match = re.search(r'(\d{6})\.pkl$', fn)
                 if match and match.group(1).startswith("25"):
                     d = int(match.group(1))
                 elif len(date_token) >= 6 and date_token.startswith("25"):
                     d = int(date_token)
        except:
            d = -1

        # 2. 옵션(h, w, n) 파싱
        h=w=n=None
        for p in parts:
            if p.startswith("h"):
                try: h=int(p[1:])
                except: pass
            
            if p.startswith("w"):
                if "full" in p.lower():
                    w = 0  # wFull은 0으로 간주
                else:
                    try: w=int(p[1:])
                    except: pass
                    
            if p.startswith("n"):
                try: n=int(p[1:])
                except: pass
        
        # 3. [핵심] 옵션 완전 일치 여부 검사 (Strict Check)
        # 하나라도 다르면 후보에서 즉시 제외
        if h == TARGET_H and w == TARGET_W and n == TARGET_N:
            # 파일 수정 시간(mtime) - 동점자 처리용
            full_path = os.path.join(real_dir, fn)
            try:
                mtime = os.path.getmtime(full_path)
            except:
                mtime = 0
            
            # (날짜, 생성시간, 파일명) 튜플 저장
            valid_candidates.append((d, mtime, fn))

    # 4. 결과 처리
    if not valid_candidates:
        # 일치하는 파일이 없으면 에러 발생 (대충 아무거나 주지 않음)
        msg = (f"❌ [Error] 조건에 맞는 엔진 파일을 찾을 수 없습니다.\n"
               f"   - 요청 조건: h={TARGET_H}, w={TARGET_W}, n={TARGET_N}\n"
               f"   - 대상 폴더: {real_dir}")
        raise FileNotFoundError(msg)

    # 5. 정렬: 날짜(d) 내림차순 -> 생성시간(mtime) 내림차순
    # 가장 최신 날짜, 가장 최근에 만들어진 파일을 0번으로 가져옴
    valid_candidates.sort(key=lambda x: (x[0], x[1]), reverse=True)
    
    best_file = valid_candidates[0][2]
    print(f"[Engine Selector] 조건 일치 파일 발견: {best_file} (Date: {valid_candidates[0][0]})")
    
    return os.path.join(real_dir, best_file)

def load_latest_db(version="V31"): # [FIXED] NameError 해결을 위해 함수 정의 복구
    """DB 디렉토리에서 최신 통합 DB 파일을 찾아 로드합니다."""
    db_dir = get_path("HOJ_DB")
    latest = find_latest_file(db_dir, f"HOJ_DB_{version}")
    if not latest:
        raise FileNotFoundError("DB를 찾지 못했습니다.")

    df = pd.read_parquet(latest)
    df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
    return df, latest


# ============================================================
# [패치 추가] 엑셀 서식 자동 조정 함수
# ============================================================
def auto_adjust_column_width(worksheet):
    """ 엑셀 컬럼 너비 자동 맞춤 및 헤더 스타일링 """
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter 
        
        # 헤더 스타일 적용
        col[0].font = header_font
        col[0].fill = header_fill
        col[0].alignment = Alignment(horizontal='center')

        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        # 너비 조정
        adjusted_width = (max_length + 2) * 1.1
        adjusted_width = min(adjusted_width, 50) # 최대 50
        adjusted_width = max(adjusted_width, 10) # 최소 10
        worksheet.column_dimensions[column].width = adjusted_width


# ============================================================
# [수정] Gemini 분석 함수 (결과 텍스트 반환하도록 변경)
# ============================================================
def get_gemini_analysis(df):
    """Gemini를 이용해 Top10 분석 텍스트를 생성하여 반환"""
    
    if not GEMINI_API_KEY:
        return "\n[Gemini] API Key가 없어 분석을 생략합니다.\n"

    try:
        genai.configure(api_key=GEMINI_API_KEY)

        # 모델 자동 선택 로직
        all_models = list(genai.list_models())
        valid_model_name = None
        
        # 1순위: Flash (빠름)
        for m in all_models:
            if 'generateContent' in m.supported_generation_methods and 'flash' in m.name:
                valid_model_name = m.name
                break
        
        # 2순위: Pro
        if valid_model_name is None:
            for m in all_models:
                if 'generateContent' in m.supported_generation_methods and 'pro' in m.name:
                    valid_model_name = m.name
                    break
        
        # 기본값
        if valid_model_name is None:
            valid_model_name = "models/gemini-1.5-flash"

        model = genai.GenerativeModel(valid_model_name)
        target_list_str = df.to_string(index=False)
        
        prompt = f"""
아래는 오늘의 HOJ Top10 종목 리스트입니다. 
이 종목들을 기반으로 상승 가능성이 높은 종목을 3개 추천해 주세요.

[Top10 종목 데이터]
{target_list_str}

[요구사항]
- 추천 사유 1~2줄 포함
- 상승 가능성이 높은 순서대로 3개만 제시

[형식]
=== Gemini's Pick ===
1. 종목명: 사유
2. 종목명: 사유
3. 종목명: 사유
"""
        print(f"[Gemini] 모델 '{valid_model_name}' 분석 실행 중...", end="\r") # 진행중 표시
        response = model.generate_content(prompt)
        print(f"[Gemini] 분석 완료.                                 ") # 지우기

        return response.text

    except Exception as e:
        return f"\n⚠ Gemini 분석 중 오류 발생: {e}\n"


# ============================================================
# 메인 로직
# ============================================================
def main(rank_by="combo", topk=10, version="V31"):

    # 1. 엔진 및 DB 로드
    eng_path = find_engine_real()
    with open(eng_path, "rb") as f:
        payload = pickle.load(f)
    model_cls = payload["model_cls"]
    model_reg = payload["model_reg"]
    features  = payload["features"]
    
    df, db_path = load_latest_db(version) # [FIXED] 이제 load_latest_db가 정의되어 오류 발생 안함
    max_date = df["Date"].max()
    df_d = df[df["Date"] == max_date].copy()
    close_col = pick_close_col(df_d)

    # 2. 피처 확인 및 예측
    db_features = [c for c in features if c in df_d.columns]
    X = df_d[db_features].copy()
    mask = X.notnull().all(axis=1)
    df_d = df_d.loc[mask].copy()
    X = X.loc[mask]

    prob = model_cls.predict_proba(X)[:,1]
    ret  = model_reg.predict(X)
    ret_clip = np.clip(ret, -0.10, None)
    combo = prob * ret_clip

    # 3. 결과 DataFrame 생성
    df_out = pd.DataFrame({
        "종목명": df_d.get("Name", df_d.get("name")),
        "종목코드": df_d.get("Code", df_d.get("code")),
        "현재가": df_d[close_col],
        "상승확률(%)": (prob*100).round(2),
        "예측수익률(%)": (ret*100).round(2),
        "동시적용 기대수익(%)": (combo*100).round(2),
    })

    # 정렬
    keymap = {"combo":"동시적용 기대수익(%)", "prob":"상승확률(%)", "ret":"예측수익률(%)"}
    sort_key = keymap.get(rank_by, "동시적용 기대수익(%)")
    df_out = df_out.sort_values(sort_key, ascending=False).head(topk)

    # ------------------------------------------------------------
    # 4. [변경] AI 분석 먼저 실행
    # ------------------------------------------------------------
    ai_result_text = get_gemini_analysis(df_out)

    # ------------------------------------------------------------
    # 5. [변경] 통합 리포트 생성 (출력 및 저장용)
    # ------------------------------------------------------------
    report_content = []
    report_content.append("=" * 60)
    report_content.append(f"📈 HOJ AI Daily Report [{max_date.date()}]")
    report_content.append("=" * 60)
    report_content.append(f"\n[1] 예측 Top {topk} (기준: {rank_by})")
    report_content.append("-" * 60)
    report_content.append(df_out.to_string(index=False))
    report_content.append("-" * 60)
    report_content.append("\n[2] Gemini AI Investment Opinion")
    report_content.append("-" * 60)
    report_content.append(ai_result_text.strip())
    report_content.append("=" * 60)
    
    full_report_str = "\n".join(report_content)

    # ------------------------------------------------------------
    # 5-1. JSON 저장 (엔진 메타 + Top10 + AI 리포트)
    # ------------------------------------------------------------
    def _json_safe(x):
        if isinstance(x, np.generic):
            return x.item()
        if isinstance(x, Path):
            return str(x)
        return x

    engine_meta = payload.get("meta", {}) if isinstance(payload, dict) else {}
    engine_info = {
        "engine_path": str(eng_path),
        "db_path": str(db_path),
        **engine_meta,
        "features": payload.get("features", []) if isinstance(payload, dict) else [],
        "feature_importances": [],
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "prediction_date": str(max_date.date()),
        "rank_by": rank_by,
        "topk": int(topk),
    }

    if payload.get("model_reg") is not None and hasattr(payload["model_reg"], "feature_importances_"):
        fi = list(zip(engine_info["features"], payload["model_reg"].feature_importances_))
        fi = sorted(fi, key=lambda x: x[1], reverse=True)
        engine_info["feature_importances"] = [
            {"name": str(n), "importance": float(v)} for n, v in fi
        ]

    top10_records = []
    for idx, row in df_out.reset_index(drop=True).iterrows():
        rec = {"rank": idx + 1}
        for col in df_out.columns:
            rec[col] = _json_safe(row[col])
        top10_records.append(rec)

    json_payload = {
        "engine_meta": engine_info,
        "top10": top10_records,
        "ai_report": ai_result_text.strip(),
        "full_report": full_report_str,
    }

    info_dir = r"F:\autostockG\MODELENGINE\INFO\hoj_engine_info"
    os.makedirs(info_dir, exist_ok=True)
    json_name = os.path.basename(eng_path).replace(".pkl", ".json")
    json_path = os.path.join(info_dir, json_name)

    # 기존 파일이 있으면 병합(메타/Top10/AI 갱신)
    if os.path.exists(json_path):
        try:
            with open(json_path, "r", encoding="utf-8") as f:
                existing = json.load(f)
            if isinstance(existing, dict):
                existing["engine_meta"] = json_payload.get("engine_meta", existing.get("engine_meta"))
                existing["top10"] = json_payload.get("top10", existing.get("top10"))
                existing["ai_report"] = json_payload.get("ai_report", existing.get("ai_report"))
                existing["full_report"] = json_payload.get("full_report", existing.get("full_report"))
                json_payload = existing
        except Exception as e:
            print(f"[WARN] 기존 JSON 로드 실패, 새로 생성합니다: {e}")

    try:
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(json_payload, f, ensure_ascii=False, indent=2)
        print(f"[SAVE]   JSON: {json_name} (엔진/Top10/AI 통합)")
    except Exception as e:
        print(f"[Error] JSON 저장 실패: {e}")

    # ------------------------------------------------------------
    # 6. 화면 출력
    # ------------------------------------------------------------
    print(full_report_str)
    print(f"\n[ENGINE] {os.path.basename(eng_path)}")
    print(f"[DB]     {os.path.basename(db_path)}")

# ============================================================
# CLI
# ============================================================
if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--rank_by", default="combo", help="combo | prob | ret")
    ap.add_argument("--topk", type=int, default=10)
    ap.add_argument("--version", default="V31")
    args = ap.parse_args()
    main(rank_by=args.rank_by, topk=args.topk, version=args.version)
