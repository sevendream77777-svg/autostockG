# ============================================================
# daily_recommender.py  (Unified: daily + predict)
# ------------------------------------------------------------
# - 자동 운용(REAL 엔진 자동선택, combo/prob/ret) 유지
# - 공용 추론 로직 흡수: load_engine, get_unified_db_path, run_prediction_core
# - 수동 엔진/날짜/TopN 모드: --engine / --date / --top
# - rank_by = combo | prob | ret | score (명확 분기)
# - Gemini AI: --ai 1 일 때만 시도. 키/라이브러리/에러 시 전부 스킵
# - JSON 저장/병합: 기본 on (--save_json 1). 타입 mismatch/필드 누락 방어.
# - DB 선택: 엔진 meta.version → 해당 version 최신 DB (find_latest_file)
# - Close 컬럼 자동 탐지: pick_close_col
# - 삭제 없음, 보강 위주(원본 로직 최대한 존중)
# ============================================================

import os, sys, argparse, pickle, warnings, json, re
from datetime import datetime
from pathlib import Path
from typing import Optional, Tuple, Dict, Any, List

import numpy as np
import pandas as pd
from textwrap import dedent

# ============ 프로젝트 경로 설정 ============
current_dir = os.path.dirname(os.path.abspath(__file__))
parent_dir  = os.path.dirname(current_dir)   # MODELENGINE
root_dir    = os.path.dirname(parent_dir)    # Root
sys.path.append(root_dir)
try:
    from MODELENGINE.UTIL.config_paths import get_path
    from MODELENGINE.UTIL.version_utils import find_latest_file
except Exception:
    sys.path.append(parent_dir)
    from UTIL.config_paths import get_path
    from UTIL.version_utils import find_latest_file

# ============ (선택) Excel 서식 ============
try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, Font, PatternFill
    _HAS_OPENPYXL = True
except Exception:
    _HAS_OPENPYXL = False

# ============ (선택) Gemini API ============
def _load_api_key() -> Optional[str]:
    key_path = r"C:\공유주방\!개인폴더\!이호정이사\각종key_appkey_decret\googlegemini_api.txt"
    try:
        if os.path.exists(key_path):
            with open(key_path, "r", encoding="utf-8") as f:
                return f.read().strip()
    except Exception:
        pass
    return None

GEMINI_API_KEY = _load_api_key()

def _safe_import_gemini():
    try:
        import google.generativeai as genai  # noqa: F401
        return True
    except Exception:
        return False

# ============================================================
# 유틸
# ============================================================
def pick_close_col(df: pd.DataFrame) -> str:
    """Close/ClosePrice/종가/가격 자동 인식"""
    cand = ["Close","close","ClosePrice","종가","가격","Adj Close","AdjClose"]
    for c in cand:
        if c in df.columns:
            return c
    nums = [c for c in df.columns if df[c].dtype.kind in ("i","f")]
    if len(nums) == 1:
        return nums[0]
    raise KeyError("종가 컬럼을 찾지 못했습니다. (Close/ClosePrice/종가/가격 등)")

def _json_safe(x):
    if isinstance(x, np.generic):
        return x.item()
    if isinstance(x, Path):
        return str(x)
    return x

def _dict_merge_safe(base: Dict[str, Any], add: Dict[str, Any]) -> Dict[str, Any]:
    """dict 병합(타입/필드 누락 방어). base를 in-place로 갱신하지 않음."""
    out = dict(base) if isinstance(base, dict) else {}
    if not isinstance(add, dict):
        return out

    for k, v in add.items():
        if k not in out:
            out[k] = v
            continue
        # 둘 다 dict면 재귀 병합, 아니면 add 우선
        if isinstance(out[k], dict) and isinstance(v, dict):
            out[k] = _dict_merge_safe(out[k], v)
        else:
            out[k] = v
    return out

# ============================================================
# 엔진 자동 선택 (Strict)
# ============================================================
def find_engine_real() -> str:
    """
    REAL 폴더에서 목표(h,w,n) 정확 일치 & 최신 날짜/mtime 우선 선택.
    필요 시 UI/CLI에서 변경 가능하도록 기본값은 여기서만 관리.
    """
    TARGET_H = 5
    TARGET_W = 60
    TARGET_N = 1000

    base_root = get_path("HOJ_ENGINE")
    if os.path.isfile(base_root):
        base_root = os.path.dirname(base_root)
    real_dir = os.path.join(base_root, "REAL")
    if not os.path.isdir(real_dir):
        raise FileNotFoundError("REAL 폴더 없음: " + real_dir)

    cands = [fn for fn in os.listdir(real_dir)
             if fn.startswith("HOJ_ENGINE_REAL") and fn.endswith(".pkl")]
    if not cands:
        raise FileNotFoundError("REAL 폴더에 엔진 파일이 없습니다.")

    def _parse(fn: str) -> Tuple[int, float, int, int, int]:
        parts = fn.split("_")
        # 날짜(YYMMDD) 파싱
        date_token = parts[-1].replace(".pkl","")
        d = -1
        try:
            if len(date_token) == 6 and date_token.startswith("25"):
                d = int(date_token)
            elif len(date_token) == 4:
                d = int("25" + date_token)
            elif len(date_token) == 8 and date_token.startswith("20"):
                d = int(date_token[2:])
            else:
                m = re.search(r'(\d{6})\.pkl$', fn)
                if m and m.group(1).startswith("25"):
                    d = int(m.group(1))
                elif len(date_token) >= 6 and date_token.startswith("25"):
                    d = int(date_token)
        except Exception:
            d = -1

        h=w=n=None
        for p in parts:
            if p.startswith("h"):
                try: h=int(p[1:])
                except: pass
            if p.startswith("w"):
                if "full" in p.lower():
                    w = 0
                else:
                    try: w=int(p[1:])
                    except: pass
            if p.startswith("n"):
                try: n=int(p[1:])
                except: pass

        mtime = 0.0
        try:
            mtime = os.path.getmtime(os.path.join(real_dir, fn))
        except Exception:
            pass
        return d, mtime, (h or -1), (w or -1), (n or -1)

    valid = []
    for fn in cands:
        d, mt, h, w, n = _parse(fn)
        if h == TARGET_H and w == TARGET_W and n == TARGET_N:
            valid.append((d, mt, fn))

    if not valid:
        raise FileNotFoundError(
            f"조건(h={TARGET_H}, w={TARGET_W}, n={TARGET_N}) 일치 파일 없음: {real_dir}"
        )

    valid.sort(key=lambda x: (x[0], x[1]), reverse=True)
    best = valid[0][2]
    print(f"[Engine Selector] 선택: {best}")
    return os.path.join(real_dir, best)

# ============================================================
# DB 선택(통일): 엔진 meta.version → 해당 version 최신 파일
# ============================================================
def get_unified_db_path(version: str) -> str:
    base = get_path("HOJ_DB")
    if os.path.isfile(base):
        base = os.path.dirname(base)

    # version prefix로 최신 파일 탐색
    prefix = f"HOJ_DB_{version}"
    latest = find_latest_file(base, prefix)
    if latest:
        return latest
    # 마지막 폴백: 고정 이름
    candidate = os.path.join(base, f"{prefix}.parquet")
    return candidate

def load_latest_db(version: str) -> Tuple[pd.DataFrame, str]:
    p = get_unified_db_path(version)
    if not os.path.exists(p):
        raise FileNotFoundError(f"DB 파일을 찾을 수 없습니다: {p}")
    df = pd.read_parquet(p)
    if "Date" in df.columns:
        df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
    return df, p

# ============================================================
# 엔진 로더 (호환/검증)
# ============================================================
def load_engine(engine_path: str) -> Dict[str, Any]:
    if not os.path.exists(engine_path):
        raise FileNotFoundError(f"엔진 파일을 찾을 수 없습니다: {engine_path}")
    with open(engine_path, "rb") as f:
        data = pickle.load(f)

    if not isinstance(data, dict):
        raise ValueError("엔진 포맷이 dict가 아닙니다.")
    if "features" not in data:
        print("⚠ 구버전 엔진: 'features' 없음. 호환성 이슈 가능.")
        data["features"] = data.get("feature_names", [])
    if "meta" not in data:
        data["meta"] = {}

    # 기본 필드 보정
    data.setdefault("model_reg", None)
    data.setdefault("model_cls", None)
    return data

# ============================================================
# 공용 추론 코어
# ============================================================
def run_prediction_core(
    engine_path: str,
    target_date: Optional[str],
    top_n: int,
    rank_by: str,
    version_override: Optional[str] = None,
) -> Tuple[pd.DataFrame, Dict[str, Any], str]:
    """
    반환: (df_out, engine_payload, db_path)
    df_out 컬럼 표준: ['Name','Code','Close','prob','ret','score','combo']
    - rank_by=score이면 score만 필수(나머지는 선택)
    """
    payload = load_engine(engine_path)
    model_reg = payload.get("model_reg")
    model_cls = payload.get("model_cls")
    features  = payload.get("features", [])
    meta      = payload.get("meta", {})
    version   = (version_override or meta.get("version") or "V31")

    # DB
    df, db_path = load_latest_db(version)
    # 날짜
    if target_date:
        td = pd.to_datetime(target_date)
    else:
        td = df["Date"].max()
    daily = df[df["Date"] == td].copy()
    if daily.empty:
        raise ValueError(f"해당 날짜({td.date()}) 데이터가 없습니다.")

    # feature 검증 & NaN 필터
    missing = [f for f in features if f not in daily.columns]
    if missing:
        raise KeyError(f"필수 피처 누락: {missing[:5]} ... ({len(missing)}개)")

    X = daily[features].copy()
    valid_mask = X.notnull().all(axis=1)
    if not valid_mask.all():
        drop_count = len(daily) - valid_mask.sum()
        print(f"⚠ 결측치로 {drop_count}개 종목 제외")
        daily = daily.loc[valid_mask].copy()
        X = X.loc[valid_mask]

    if len(daily) == 0:
        raise ValueError("예측 가능한 종목이 없습니다(전체 결측).")

    # 컬럼 표준화용 보조
    name_col = "Name" if "Name" in daily.columns else ("name" if "name" in daily.columns else None)
    code_col = "Code" if "Code" in daily.columns else ("code" if "code" in daily.columns else None)
    close_col = pick_close_col(daily)

    # 예측
    df_out = pd.DataFrame()
    df_out["Name"] = daily[name_col] if name_col else ""
    df_out["Code"] = daily[code_col] if code_col else ""
    df_out["Close"] = daily[close_col]

    # rank_by 분기
    rank_by = (rank_by or "combo").lower()
    score = None; prob = None; ret = None; combo = None

    # score(회귀)만 필요한 경우: 최소 계산
    if rank_by == "score":
        if model_reg is None:
            raise ValueError("rank_by=score인데 회귀 모델이 없습니다(model_reg=None).")
        score = model_reg.predict(X)
        df_out["score"] = score.astype(float)
        # 선택적으로 prob/ret도 담고 싶다면 여기서 계산 가능(비용 고려)
        if model_cls is not None:
            prob = model_cls.predict_proba(X)[:, 1]
            df_out["prob"] = prob.astype(float)
        else:
            df_out["prob"] = 0.0
        # ret은 score와 동일 개념일 수 있어 중복 계산 생략
        df_out["ret"] = score.astype(float)
        df_out["combo"] = (df_out["prob"] * np.clip(df_out["ret"], -0.10, None)).astype(float)

        df_out = df_out.sort_values("score", ascending=False).head(top_n)
        return df_out.reset_index(drop=True), payload, db_path

    # combo/prob/ret: 기존 daily 방식
    # 1) prob
    if model_cls is not None:
        prob = model_cls.predict_proba(X)[:, 1].astype(float)
    else:
        prob = np.zeros(len(X), dtype=float)

    # 2) ret (회귀)
    if model_reg is not None:
        ret = model_reg.predict(X).astype(float)
    else:
        ret = np.zeros(len(X), dtype=float)

    # 3) score = ret 로 통일(표준 컬럼 유지 목적)
    score = ret.copy()
    # 4) combo = prob * ret_clip
    ret_clip = np.clip(ret, -0.10, None)
    combo = prob * ret_clip

    df_out["prob"]  = prob
    df_out["ret"]   = ret
    df_out["score"] = score
    df_out["combo"] = combo

    # 정렬키
    keymap = {"combo":"combo", "prob":"prob", "ret":"ret"}
    sort_key = keymap.get(rank_by, "combo")
    df_out = df_out.sort_values(sort_key, ascending=False).head(top_n)
    return df_out.reset_index(drop=True), payload, db_path

# ============================================================
# Gemini 분석 (옵션)
# ============================================================
def get_gemini_analysis(df_out: pd.DataFrame, do_ai: bool) -> str:
    if not do_ai:
        return "[AI] 옵션 비활성화로 분석 생략"
    if not _safe_import_gemini():
        return "[AI] 라이브러리 미설치로 분석 생략"
    if not GEMINI_API_KEY:
        return "[AI] API Key 미존재로 분석 생략"
    try:
        import google.generativeai as genai
        genai.configure(api_key=GEMINI_API_KEY)

        # 모델 선택
        valid_model_name = "models/gemini-1.5-flash"
        try:
            models = list(genai.list_models())
            for m in models:
                if 'generateContent' in getattr(m, "supported_generation_methods", []) and 'flash' in m.name:
                    valid_model_name = m.name; break
        except Exception:
            pass

        model = genai.GenerativeModel(valid_model_name)
        prompt = f"""아래는 오늘의 Top 리스트입니다. 상승 가능성 높은 3개와 간단 사유를 주세요.
[Top]
{df_out.to_string(index=False)}
형식:
1) 종목명: 사유
2) 종목명: 사유
3) 종목명: 사유
"""
        resp = model.generate_content(prompt)
        return resp.text or "[AI] 응답 비어있음"
    except Exception as e:
        return f"[AI] 오류로 분석 생략: {e}"

# ============================================================
# JSON 저장/병합
# ============================================================
def save_json_payload(
    engine_path: str,
    db_path: str,
    payload: Dict[str, Any],
    df_out: pd.DataFrame,
    ai_text: str,
    rank_by: str,
    top_n: int,
    prediction_date: datetime,
    save_json: bool = True,
) -> Optional[str]:
    if not save_json:
        return None

    info_dir = r"F:\autostockG\MODELENGINE\INFO\hoj_engine_info"
    os.makedirs(info_dir, exist_ok=True)
    json_name = os.path.basename(engine_path).replace(".pkl", ".json")
    json_path = os.path.join(info_dir, json_name)

    # 메타
    engine_meta = dict(payload.get("meta", {}))
    features = payload.get("features", [])
    # feature_importances (옵션)
    feat_imps = []
    mr = payload.get("model_reg")
    if mr is not None and hasattr(mr, "feature_importances_"):
        try:
            fi = list(zip(features, mr.feature_importances_))
            fi = sorted(fi, key=lambda x: x[1], reverse=True)
            feat_imps = [{"name": str(n), "importance": float(v)} for n, v in fi]
        except Exception:
            feat_imps = []

    # TopN 레코드
    top_records = []
    for i, row in df_out.reset_index(drop=True).iterrows():
        rec = {"rank": int(i + 1)}
        for c in ["Name","Code","Close","prob","ret","score","combo"]:
            if c in df_out.columns:
                rec[c] = _json_safe(row[c])
        top_records.append(rec)

    engine_info = {
        "engine_path": str(engine_path),
        "db_path": str(db_path),
        **engine_meta,
        "features": features,
        "feature_importances": feat_imps,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "prediction_date": prediction_date.strftime("%Y-%m-%d"),
        "rank_by": rank_by,
        "topk": int(top_n),
    }

    # full_report 간단 버전
    header = f"📈 HOJ AI Daily Report [{prediction_date.strftime('%Y-%m-%d')}]"
    body = df_out.to_string(index=False)
    full_report = "\n".join([
        "="*60, header, "="*60,
        f"[Top {top_n}] (rank_by={rank_by})",
        body,
        "-"*60,
        "[AI]",
        ai_text or "",
        "="*60
    ])

    new_payload = {
        "engine_meta": engine_info,
        "top10": top_records,
        "ai_report": ai_text or "",
        "full_report": full_report,
    }

    # 기존 파일 병합
    if os.path.exists(json_path):
        try:
            with open(json_path, "r", encoding="utf-8") as f:
                old = json.load(f)
            if not isinstance(old, dict):
                old = {}
        except Exception:
            old = {}
        merged = _dict_merge_safe(old, new_payload)
        data_to_save = merged
    else:
        data_to_save = new_payload

    try:
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(data_to_save, f, ensure_ascii=False, indent=2)
        print(f"[SAVE] JSON: {json_name}")
        return json_path
    except Exception as e:
        print(f"[ERROR] JSON 저장 실패: {e}")
        return None

# ============================================================
# 메인
# ============================================================
def main(
    rank_by: str = "combo",
    topk: int = 10,
    version: str = "V31",
    engine_path: Optional[str] = None,
    date_str: Optional[str] = None,
    ai_flag: int = 0,
    save_json_flag: int = 1,
):
    """
    동작 규칙(요약):
    - 자동탐색: --engine 미지정 → REAL 폴더에서 (h,w,n) 일치 최신 엔진 자동 선택
    - 수동엔진: --engine 지정 → 자동탐색 비활성, 해당 엔진 강제 사용
    - 날짜: --date 지정 시 해당 날짜, 미지정 시 DB 최신 날짜
    - DB: 엔진 meta.version → 그 버전의 최신 DB 자동 선택
    - AI: --ai 1 인 경우에만 시도, 실패/무설치/키없음 전부 우아하게 스킵
    - JSON: --save_json 1(기본) 저장/병합, 0이면 미저장
    """
    # 1) 엔진 경로 확정
    if engine_path:
        eng = engine_path
        print("[Mode] 수동 엔진 모드")
    else:
        print("[Mode] 자동탐색 모드")
        eng = find_engine_real()

    # 2) 추론 실행
    df_out, payload, db_path = run_prediction_core(
        engine_path=eng,
        target_date=date_str,
        top_n=topk,
        rank_by=rank_by,
        version_override=version,
    )

    # 3) AI 분석 (옵션)
    ai_text = get_gemini_analysis(df_out, do_ai=bool(ai_flag))

    # 4) 저장/출력
    # prediction_date
    if date_str:
        pred_dt = pd.to_datetime(date_str)
    else:
        # DB에서 도출된 날짜(= run_prediction_core 내부 선택) 사용
        # df_out에는 날짜 컬럼이 없으므로, 파일명/경로에서 유추하지 말고
        # load에서 쓴 td를 반환하지 않는 구조이므로 재계산
        # -> db_path에서 읽어 다시 max(Date)로 맞춤 (일관성 유지)
        df_tmp = pd.read_parquet(db_path)
        df_tmp["Date"] = pd.to_datetime(df_tmp["Date"], errors="coerce")
        pred_dt = df_tmp["Date"].max()

    json_out = save_json_payload(
        engine_path=eng,
        db_path=db_path,
        payload=payload,
        df_out=df_out,
        ai_text=ai_text,
        rank_by=rank_by,
        top_n=topk,
        prediction_date=pred_dt,
        save_json=bool(save_json_flag),
    )

    # 화면 출력
    print("="*60)
    print(f"[Top {topk}] (rank_by={rank_by})")
    print(df_out.to_string(index=False))
    print("-"*60)
    print(ai_text)
    print("="*60)
    print(f"[ENGINE] {os.path.basename(eng)}")
    print(f"[DB]     {os.path.basename(db_path)}")
    if json_out:
        print(f"[JSON]   {json_out}")

# ============================================================
# CLI
# ============================================================
if __name__ == "__main__":
    ap = argparse.ArgumentParser(
        prog="daily_recommender.py",
        description=dedent(\"\"\"
        HOJ 일일 추천/예측 통합 엔트리

        자동탐색 모드: --engine 미지정 → REAL 폴더에서 (h,w,n) 정확 일치 최신 엔진 자동 선택
        수동엔진 모드: --engine 지정 → 자동탐색 비활성, 해당 경로 엔진 강제 사용
        날짜 선택    : --date 지정 시 해당 날짜, 미지정 시 DB 최신 날짜
        DB 선택     : 엔진 meta.version → 해당 version 최신 DB 자동 선택
        AI          : --ai 1 일 때만 실행. 키 없음/미설치/오류는 모두 우아하게 스킵
        JSON 저장   : --save_json 1(기본) 저장/병합, 0이면 저장 안 함
        rank_by     : combo | prob | ret | score
        \"\"\"
        ).strip()
    )
    ap.add_argument("--rank_by", default="combo", help="combo | prob | ret | score")
    ap.add_argument("--topk", type=int, default=10, help="TopK (daily 스타일)")
    ap.add_argument("--version", default="V31", help="엔진 메타에 version 없을 때만 폴백")
    ap.add_argument("--engine", type=str, default=None, help="수동 엔진 경로(.pkl). 주면 자동탐색 비활성")
    ap.add_argument("--date", type=str, default=None, help="YYYY-MM-DD. 미지정 시 DB 최신 날짜 사용")
    ap.add_argument("--ai", type=int, default=0, help="1=Gemini 실행, 0=skip")
    ap.add_argument("--save_json", type=int, default=1, help="1=JSON 저장/병합, 0=미저장")

    args = ap.parse_args()
    main(
        rank_by=args.rank_by,
        topk=args.topk,
        version=args.version,
        engine_path=args.engine,
        date_str=args.date,
        ai_flag=args.ai,
        save_json_flag=args.save_json,
    )
