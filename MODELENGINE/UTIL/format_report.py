import pandas as pd
import os
import glob
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill

# ============================================================
# 설정: 파일이 있는 폴더 (현재 폴더 혹은 지정된 경로)
# ============================================================
# daily_recommender.py가 결과를 저장하는 OUTPUT 폴더 경로를 찾거나 현재 폴더를 뒤집니다.
SEARCH_DIR = os.getcwd() 
OUTPUT_FILE_NAME = 'Final_Formatted_Report.xlsx'

def find_latest_files():
    # CSV (Top 10 데이터) 찾기
    csv_files = glob.glob(os.path.join(SEARCH_DIR, '**', 'recommendation_HOJ_*.csv'), recursive=True)
    # TXT (AI 리포트) 찾기
    txt_files = glob.glob(os.path.join(SEARCH_DIR, '**', 'Report_HOJ_*.txt'), recursive=True)
    
    if not csv_files:
        print('[Error] recommendation CSV 파일을 찾을 수 없습니다. (daily_recommender.py를 먼저 실행했나요?)')
        return None, None
        
    # 가장 최근 파일 선택
    latest_csv = max(csv_files, key=os.path.getctime)
    
    # CSV와 짝이 맞는 TXT 찾기 (없으면 가장 최근 TXT 사용)
    latest_txt = None
    if txt_files:
        # 같은 타임스탬프를 가진 파일 우선 검색
        timestamp = latest_csv.split('_')[-2] # 파일명 규칙에 따라 변경 가능
        matched_txt = [f for f in txt_files if timestamp in f]
        if matched_txt:
            latest_txt = matched_txt[0]
        else:
            latest_txt = max(txt_files, key=os.path.getctime)
            
    return latest_csv, latest_txt

def auto_adjust_column_width(worksheet):
    """ 엑셀 컬럼 너비 자동 맞춤 및 헤더 스타일링 """
    # 헤더 스타일 정의 (진하게, 배경색)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        
        # 헤더 스타일 적용
        col[0].font = header_font
        col[0].fill = header_fill
        col[0].alignment = Alignment(horizontal='center')

        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        # 너비 조정 (너무 좁거나 넓지 않게 제한)
        adjusted_width = (max_length + 2) * 1.1
        adjusted_width = min(adjusted_width, 50) # 최대 50
        adjusted_width = max(adjusted_width, 10) # 최소 10
        worksheet.column_dimensions[column].width = adjusted_width

def main():
    print('\n' + '='*60)
    print('   [HOJ ENGINE] 결과 보고서 엑셀 변환기 (Excel Formatter)')
    print('='*60 + '\n')
    
    csv_path, txt_path = find_latest_files()
    
    if not csv_path:
        return

    print(f'[*] Found Data (Top 10): {os.path.basename(csv_path)}')
    if txt_path:
        print(f'[*] Found Report (AI)  : {os.path.basename(txt_path)}')
    else:
        print('[!] AI Report TXT 파일을 찾지 못했습니다. (데이터만 변환합니다)')

    try:
        # 1. 데이터 로드
        df = pd.read_csv(csv_path)
        
        # 2. AI 리포트 읽기
        ai_report = ""
        if txt_path:
            with open(txt_path, 'r', encoding='utf-8') as f:
                ai_report = f.read()

        # 3. 엑셀 저장 (Pandas)
        print(f'\n[*] Generating Excel: {OUTPUT_FILE_NAME}...')
        with pd.ExcelWriter(OUTPUT_FILE_NAME, engine='openpyxl') as writer:
            # Sheet 1: Top 10 Data
            df.to_excel(writer, sheet_name='Top 10 추천', index=False)
            
            # Sheet 2: AI Analysis (텍스트가 있는 경우)
            if ai_report:
                # 데이터프레임 없이 워크시트만 추가하는 방식이 복잡하므로
                # 텍스트를 담은 DataFrame 생성
                df_report = pd.DataFrame({'AI 분석 리포트': [ai_report]})
                df_report.to_excel(writer, sheet_name='AI 해석', index=False)

        # 4. 엑셀 서식 다듬기 (OpenPyXL)
        wb = load_workbook(OUTPUT_FILE_NAME)
        
        # (1) Top 10 시트 서식
        if 'Top 10 추천' in wb.sheetnames:
            ws = wb['Top 10 추천']
            auto_adjust_column_width(ws)
            
        # (2) AI 해석 시트 서식
        if 'AI 해석' in wb.sheetnames:
            ws = wb['AI 해석']
            cell = ws['A2'] # 본문 셀
            cell.alignment = Alignment(wrap_text=True, vertical='top') # 줄바꿈 허용
            ws.column_dimensions['A'].width = 100 # 넓게 잡기
            
            # 행 높이 자동 조절이 어려우므로 대략적으로 늘림 (내용 길이에 비례)
            line_count = ai_report.count('\n') + (len(ai_report) // 100)
            ws.row_dimensions[2].height = max(line_count * 15, 400) # 적당히 높이 조절

        wb.save(OUTPUT_FILE_NAME)
        print(f'\n[Success] 변환 완료! 파일을 확인하세요: {os.path.abspath(OUTPUT_FILE_NAME)}')
        
    except Exception as e:
        print(f'\n[Error] 변환 중 오류 발생: {e}')

if __name__ == '__main__':
    main()
